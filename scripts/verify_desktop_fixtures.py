#!/usr/bin/env python3
"""
verify_desktop_fixtures.py — Dependency-light fixture integrity check.

Purpose
-------
Independently verify that the synthetic Desktop-Local fixtures still contain
the DATA that the SQLite workflow smoke tests (B1-B5 + dedup) rely on.
This is a *fixture drift guard* — it does NOT run the real pytest suite and
does NOT replace it. The D3 gate still requires:
    cd backend && pytest tests/test_desktop_sqlite_workflow.py -v -p no:randomly

Why this exists
---------------
The real suite needs sqlalchemy/pydantic/fastapi/pytest. In some restricted
environments those cannot be installed. This script needs only the Python
stdlib plus openpyxl, so the fixture preconditions can always be checked.

Run
---
    python scripts/verify_desktop_fixtures.py
Exit code 0 = all fixture preconditions satisfied; 1 = at least one failed.

Safety: reads only synthetic fixtures. No patient data, no DB writes.
"""
from __future__ import annotations

import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except Exception as exc:  # noqa: BLE001
    print(f"FATAL: openpyxl is required: {exc}")
    sys.exit(2)

ROOT = Path(__file__).resolve().parent.parent
FIX = ROOT / "tests" / "fixtures" / "desktop_local"
MULTISHEET = FIX / "target_group_multisheet.xlsx"
CID_CONST = FIX / "cid_constants.py"

CERVICAL_LABEL = "ตรวจคัดกรองมะเร็งปากมดลูก"

results: list[tuple[bool, str]] = []


def check(cond: bool, msg: str) -> None:
    results.append((bool(cond), msg))


def thai_cid_valid(cid: str) -> bool:
    if not cid or len(cid) != 13 or not cid.isdigit():
        return False
    s = sum(int(cid[i]) * (13 - i) for i in range(12))
    return (11 - (s % 11)) % 10 == int(cid[12])


def load_cids() -> dict[str, str]:
    ns: dict[str, str] = {}
    for line in CID_CONST.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("CID_") or line.startswith(("INVALID_CID", "MISSING_CID")):
            name, _, val = line.partition("=")
            ns[name.strip()] = val.strip().strip('"').strip("'")
    return ns


def main() -> int:
    check(MULTISHEET.exists(), f"multisheet fixture exists: {MULTISHEET.name}")
    check(CID_CONST.exists(), f"cid_constants.py exists")
    if not MULTISHEET.exists() or not CID_CONST.exists():
        report()
        return 1

    cids = load_cids()
    for key in ("CID_ALICE", "CID_BOB", "CID_CHARLIE", "CID_DAVE", "CID_EVE", "INVALID_CID"):
        check(key in cids, f"constant defined: {key}")

    # Checksums: 5 synthetic CIDs valid, INVALID_CID invalid (B1 relies on this)
    for key in ("CID_ALICE", "CID_BOB", "CID_CHARLIE", "CID_DAVE", "CID_EVE"):
        check(thai_cid_valid(cids.get(key, "")), f"{key} passes Thai mod-11 checksum")
    check(not thai_cid_valid(cids.get("INVALID_CID", "")), "INVALID_CID fails checksum (B1)")
    check(cids.get("MISSING_CID", "x") == "", "MISSING_CID is empty string (B2)")

    wb = load_workbook(MULTISHEET, data_only=True)
    names = wb.sheetnames
    # Rule 2/4: must read every sheet — fixture must have roster + history sheets
    check(len(names) >= 2, f"multi-sheet workbook (>=2 sheets): {names}")

    # Identify sheets by header content (not by assuming order)
    roster = history = None
    for ws in wb.worksheets:
        header = [str(c.value).strip() if c.value is not None else "" for c in ws[1]]
        if "ชื่อบริการ" in header and "วันที่ตรวจ" in header:
            history = ws
        elif "CID" in header or "ชื่อ-สกุล" in header:
            if roster is None:
                roster = ws
    check(history is not None, "history sheet found via header 'ชื่อบริการ'+'วันที่ตรวจ' (D2.15 fix)")
    check(roster is not None, "roster sheet found")

    if history is not None:
        hdr = [str(c.value).strip() if c.value is not None else "" for c in history[1]]
        # D2.15 fix: column must be 'ชื่อบริการ', NOT 'ประเภทบริการ'
        check("ชื่อบริการ" in hdr, "history service column is 'ชื่อบริการ'")
        check("ประเภทบริการ" not in hdr, "old unmapped column 'ประเภทบริการ' is NOT present")

        rows = list(history.iter_rows(min_row=2, values_only=True))
        col = {h: i for i, h in enumerate(hdr)}
        ci, si, di = col.get("CID"), col.get("ชื่อบริการ"), col.get("วันที่ตรวจ")

        def hist_for(cid: str):
            out = []
            for r in rows:
                if r[ci] is not None and str(r[ci]).strip() == cid:
                    out.append((str(r[si]).strip() if r[si] else "", str(r[di]).strip() if r[di] else ""))
            return out

        bob = hist_for(cids["CID_BOB"])
        check(len(bob) >= 1 and all(s == CERVICAL_LABEL for s, _ in bob),
              f"B4: BOB has cervical history in TG file only ({len(bob)} rows)")
        eve = hist_for(cids["CID_EVE"])
        eve_cervical_dates = [d for s, d in eve if s == CERVICAL_LABEL]
        check("01/05/2022" in eve_cervical_dates,
              f"B5: EVE cervical date 01/05/2022 present in TG history (got {eve_cervical_dates})")
        check("20/12/2023" not in eve_cervical_dates and "2023-12-20" not in eve_cervical_dates,
              "B5: EVE cervical date is NOT the diabetes date 2023-12-20")
        dave_hist = hist_for(cids["CID_DAVE"])
        check(len(dave_hist) >= 1, "B3: DAVE present in history sheet")

    if roster is not None:
        rhdr = [str(c.value).strip() if c.value is not None else "" for c in roster[1]]
        rcol = {h: i for i, h in enumerate(rhdr)}
        rci = rcol.get("CID")
        rrows = list(roster.iter_rows(min_row=2, values_only=True))
        roster_cids = [str(r[rci]).strip() if r[rci] is not None else "" for r in rrows]
        check(cids["CID_DAVE"] in roster_cids, "B3: DAVE present in roster sheet (=> dedup to 1 row)")
        check(cids["INVALID_CID"] in roster_cids, "B1: invalid-checksum CID staged in roster")
        check(roster_cids.count("") >= 1, "B2: at least one blank-CID row staged in roster")

    return report()


def report() -> int:
    passed = sum(1 for ok, _ in results if ok)
    failed = len(results) - passed
    print("=" * 64)
    print("Desktop fixture integrity check (NOT a substitute for pytest)")
    print("=" * 64)
    for ok, msg in results:
        print(f"  [{'PASS' if ok else 'FAIL'}] {msg}")
    print("-" * 64)
    print(f"  total={len(results)}  passed={passed}  failed={failed}")
    print("=" * 64)
    if failed == 0:
        print("Fixture data supports B1-B5 + dedup. Still run the real pytest suite for the D3 gate.")
    else:
        print("Fixture drift detected — fix fixtures before trusting the smoke suite.")
    return 0 if failed == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
